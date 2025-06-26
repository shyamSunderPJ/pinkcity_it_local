# Copyright (c) 2025, pinkcity and contributors
# For license information, please see license.txt

import frappe, openpyxl, json, io
import openpyxl.styles
import openpyxl.workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from frappe.utils import flt


def execute(filters):
    # headings = get_heading()
    columns = get_columns()
    data = get_data(filters)
    # hdfc_total = sum(tss.bank_name_new == "HDFC")
    # other_total = sum(tss.bank_name_new != "HDFC")
    # grand_total = hdfc_total + other_total
    # data.append({
	# 	"basic": "<p style='font-weight: bold;'>HDFC BANK</p>",
	# 	"washing_allowance": round(hdfc_total, 2)
	# })
    # data.append({
	# 	"basic": "<p style='font-weight: bold;'>OTHER BANK</p>",
	# 	"washing_allowance": round(other_total, 2)
	# })
    # data.append({
	# 	"basic": "<p style='font-weight: bold;'>GRAND TOTAL</p>",
	# 	"washing_allowance": round(grand_total, 2)
	# })
    return columns, data



# def get_data(filters):
# 	conditions = get_conditions(filters)

# 	query = f"""
# 				SELECT 
# 					tss.occupation,
# 					tss.company,
# 					SUM(tss.gross_pay) AS gross,
# 					COUNT(tss.employee_name) AS employee,
# 					SUM(tsd.amount) AS basic,
# 					SUM(CASE WHEN tsd.salary_component = 'City Compensatory Allowance' THEN tsd.amount ELSE 0 END) AS city_compensatory_allowance,
# 					SUM(CASE WHEN tsd.salary_component = 'Washing Allowance' THEN tsd.amount ELSE 0 END) AS washing_allowance,
# 					SUM(CASE WHEN tsd.salary_component = 'House Rent Allowance' THEN tsd.amount ELSE 0 END) AS house_rent_allowance,
# 					SUM(CASE WHEN tsd.salary_component = 'OT' THEN tsd.amount ELSE 0 END) AS overtime,
# 					SUM(CASE WHEN tsd.salary_component = 'Arrear' THEN tsd.amount ELSE 0 END) AS arrear,
# 					SUM(CASE WHEN tsd.salary_component = 'Incentive Pay' THEN tsd.amount ELSE 0 END) AS incentive_pay,
# 					SUM(CASE WHEN tsd.salary_component = 'Leave Encashment' THEN tsd.amount ELSE 0 END) AS leave_encashment,
# 					SUM(CASE WHEN tsd.salary_component = 'Travelling Allowance' THEN tsd.amount ELSE 0 END) AS travelling_allowance,
# 					SUM(CASE WHEN tsd.salary_component = 'ABRY Scheme' THEN tsd.amount ELSE 0 END) AS abry_scheme,
# 					SUM(CASE WHEN tsd.salary_component = 'Earned Gross' THEN tsd.amount ELSE 0 END) AS earned_gross,
# 					SUM(CASE WHEN tsd.salary_component = 'Net Gross' THEN tsd.amount ELSE 0 END) AS net_gross,
# 					SUM(CASE WHEN tsd.salary_component = 'Employee Contribution PF' THEN tsd.amount ELSE 0 END) AS employee_contribution_pf,
# 					SUM(CASE WHEN tsd.salary_component = 'Employee Contribution ESI' THEN tsd.amount ELSE 0 END) AS employee_contribution_esi,
# 					SUM(CASE WHEN tsd.salary_component = 'Late Hour Deduction' THEN tsd.amount ELSE 0 END) AS late_hours_deduction,
# 					SUM(CASE WHEN tsd.salary_component = 'Advance Deduction' THEN tsd.amount ELSE 0 END) AS advance_deduction,
# 					SUM(CASE WHEN tsd.salary_component = 'Tax Deducted at Source' THEN tsd.amount ELSE 0 END) AS tax_deducted_at_source,
# 					SUM(CASE WHEN tsd.salary_component = 'Professional Tax' THEN tsd.amount ELSE 0 END) AS professional_tax,
# 					SUM(CASE WHEN tsd.salary_component = 'Income Tax' THEN tsd.amount ELSE 0 END) AS income_tax,
# 					SUM(CASE WHEN tsd.salary_component = 'GMI' THEN tsd.amount ELSE 0 END) AS gmi,
# 					SUM(CASE WHEN tsd.salary_component = 'Total' THEN tsd.amount ELSE 0 END) AS total
					
# 				FROM `tabSalary Component` AS tsc
# 				LEFT JOIN `tabSalary Detail` AS tsd 
# 					ON tsd.abbr = tsc.salary_component_abbr
# 				LEFT JOIN `tabSalary Slip` AS tss 
# 					ON tss.name = tsd.parent
# 				WHERE tss.occupation IN ('Worker', 'Staff')  
# 				AND tss.docstatus <= 1                     
# 				GROUP BY 
# 					tss.occupation
# 				ORDER BY 
# 					tss.occupation;


# 				{conditions}
# 				"""


# 	return frappe.db.sql(query, as_dict=1,)



def get_columns():
	return [
        {"fieldname": "sr_no", "fieldtype": "Data", "label": "S.No", "width": 120},
        {"fieldname": "occupation", "fieldtype": "Data", "label": "Occupation", "width": 120},
        {"fieldname": "gross", "fieldtype": "Currency", "label": "Gross Pay", "width": 120},
        {"fieldname": "employee", "fieldtype": "Data", "label": "Employee", "options": "Employee", "width": 120},
        {"fieldname": "basic", "fieldtype": "Currency", "label": "Basic", "width": 120, "group": "Earnings"},
        {"fieldname": "city_compensatory_allowance", "fieldtype": "Currency", "label": "City Compensatory Allowance", "width": 150, "group": "Earnings"},
        {"fieldname": "washing_allowance", "fieldtype": "Currency", "label": "Washing Allowance", "width": 120, "group": "Earnings"},
        {"fieldname": "house_rent_allowance", "fieldtype": "Currency", "label": "House Rent Allowance", "width": 150, "group": "Earnings"},
        {"fieldname": "overtime", "fieldtype": "Currency", "label": "Overtime", "width": 120, "group": "Earnings"},
        {"fieldname": "arrear", "fieldtype": "Currency", "label": "Arrear", "width": 120, "group": "Earnings"},
        {"fieldname": "incentive_pay", "fieldtype": "Currency", "label": "Incentive Pay", "width": 120, "group": "Earnings"},
        {"fieldname": "leave_encashment", "fieldtype": "Currency", "label": "Leave Encashment", "width": 140, "group": "Earnings"},
        {"fieldname": "travelling_allowance", "fieldtype": "Currency", "label": "Travelling Allowance", "width": 140, "group": "Earnings"},
        {"fieldname": "abry_scheme", "fieldtype": "Currency", "label": "Abry Scheme", "width": 120, "group": "Earnings"},
        {"fieldname": "earned_gross", "fieldtype": "Currency", "label": "Earned Gross", "width": 120, "group": "Earnings"},
        {"fieldname": "net_gross", "fieldtype": "Currency", "label": "Net Gross", "width": 120, "group": "Earnings"},
        {"fieldname": "employee_contribution_pf", "fieldtype": "Currency", "label": "Employee Contribution Pf", "width": 160, "group": "Deductions"},
        {"fieldname": "employee_contribution_esi", "fieldtype": "Currency", "label": "Employee Contribution Esi", "width": 160, "group": "Deductions"},
        {"fieldname": "late_hours_deduction", "fieldtype": "Currency", "label": "Late Hours Deduction", "width": 140, "group": "Deductions"},
        {"fieldname": "advance_deduction", "fieldtype": "Currency", "label": "Advance Deduction", "width": 140, "group": "Deductions"},
        {"fieldname": "tax_deducted_at_source", "fieldtype": "Currency", "label": "Tax Deducted At Source", "width": 160, "group": "Deductions"},
        {"fieldname": "professional_tax", "fieldtype": "Currency", "label": "Professional Tax", "width": 140, "group": "Deductions"},
        {"fieldname": "income_tax", "fieldtype": "Currency", "label": "Income Tax", "width": 120, "group": "Deductions"},
        {"fieldname": "gmi", "fieldtype": "Currency", "label": "Gmi", "width": 100, "group": "Deductions"},
        {"fieldname": "total", "fieldtype": "Currency", "label": "Total", "width": 120, "group": "Deductions"}
    ]

# def get_conditions(filters):
# 	conditions = ""
	
# 	if filters.get("company"):
# 		conditions += " AND tss.company = '" + filters.get("company") +"' " 

# 	if filters.get("month"):
# 		month = [
# 			"January",
# 			"Febuary",
# 			"March",
# 			"April",
# 			"May",
# 			"June",
# 			"July",
# 			"August",
# 			"September",
# 			"October",
# 			"November",
# 			"December",
# 		].index(filters["month"]) + 1
# 		conditions += f" AND MONTH(tss.posting_date) = {month} "
	
# 	if filters.get("year"):
# 		conditions += f" AND YEAR(tss.posting_date) =  " + filters.get("year") +" " 
	
# 	if filters.get("employee"):
# 		conditions += f" AND tss.employee =  '" + filters.get("employee") +"' " 
	

# 	return conditions


def get_data(filters):
    conditions = get_conditions(filters)

    # Create the query with conditions directly embedded
    query = f"""
        SELECT 
            tss.occupation,
            tss.company,
            SUM(tss.gross_monthly_salary) AS gross,
            COUNT(tss.employee) AS employee,
            ( SELECT SUM(tsd.amount) FROM `tabSalary Detail` AS tsd WHERE tsd.parent = tss.name AND tsd.abbr = 'B' ) basic
        FROM `tabSalary Slip` AS tss 
        WHERE tss.docstatus <= 1
        {conditions}
        GROUP BY 
            tss.occupation
        ORDER BY 
            tss.occupation;
    """
    
	# ,
    #         SUM(tsd.amount) AS basic,
    #         SUM(CASE WHEN tsd.salary_component = 'City Compensatory Allowance' THEN tsd.amount ELSE 0 END) AS city_compensatory_allowance,
    #         SUM(CASE WHEN tsd.salary_component = 'Washing Allowance' THEN tsd.amount ELSE 0 END) AS washing_allowance,
    #         SUM(CASE WHEN tsd.salary_component = 'House Rent Allowance' THEN tsd.amount ELSE 0 END) AS house_rent_allowance,
    #         SUM(CASE WHEN tsd.salary_component = 'OT' THEN tsd.amount ELSE 0 END) AS overtime,
    #         SUM(CASE WHEN tsd.salary_component = 'Arrear' THEN tsd.amount ELSE 0 END) AS arrear,
    #         SUM(CASE WHEN tsd.salary_component = 'Incentive Pay' THEN tsd.amount ELSE 0 END) AS incentive_pay,
    #         SUM(CASE WHEN tsd.salary_component = 'Leave Encashment' THEN tsd.amount ELSE 0 END) AS leave_encashment,
    #         SUM(CASE WHEN tsd.salary_component = 'Travelling Allowance' THEN tsd.amount ELSE 0 END) AS travelling_allowance,
    #         SUM(CASE WHEN tsd.salary_component = 'ABRY Scheme' THEN tsd.amount ELSE 0 END) AS abry_scheme,
    #         SUM(CASE WHEN tsd.salary_component = 'Earned Gross' THEN tsd.amount ELSE 0 END) AS earned_gross,
    #         SUM(CASE WHEN tsd.salary_component = 'Net Gross' THEN tsd.amount ELSE 0 END) AS net_gross,
    #         SUM(CASE WHEN tsd.salary_component = 'Employee Contribution PF' THEN tsd.amount ELSE 0 END) AS employee_contribution_pf,
    #         SUM(CASE WHEN tsd.salary_component = 'Employee Contribution ESI' THEN tsd.amount ELSE 0 END) AS employee_contribution_esi,
    #         SUM(CASE WHEN tsd.salary_component = 'Late Hour Deduction' THEN tsd.amount ELSE 0 END) AS late_hours_deduction,
    #         SUM(CASE WHEN tsd.salary_component = 'Advance Deduction' THEN tsd.amount ELSE 0 END) AS advance_deduction,
    #         SUM(CASE WHEN tsd.salary_component = 'Tax Deducted at Source' THEN tsd.amount ELSE 0 END) AS tax_deducted_at_source,
    #         SUM(CASE WHEN tsd.salary_component = 'Professional Tax' THEN tsd.amount ELSE 0 END) AS professional_tax,
    #         SUM(CASE WHEN tsd.salary_component = 'Income Tax' THEN tsd.amount ELSE 0 END) AS income_tax,
    #         SUM(CASE WHEN tsd.salary_component = 'GMI' THEN tsd.amount ELSE 0 END) AS gmi,
    #         SUM(CASE WHEN tsd.salary_component = 'Total' THEN tsd.amount ELSE 0 END) AS total

	# FROM `tabSalary Detail` AS tsd 
    #     LEFT JOIN `tabSalary Slip` AS tss ON tss.name = tsd.parent
    
    # Execute the query directly without using params
    return frappe.db.sql(query, as_dict=1)

def get_conditions(filters):
    conditions = ""
    
    if filters.get("company"):
        conditions += " AND tss.company = '{}'".format(filters.get("company"))

    if filters.get("month"):
        month = [
            "January", 
            "February",
            "March",
            "April",
            "May", 
            "June", 
            "July", 
            "August", 
            "September", 
            "October", 
            "November", 
            "December"
        ].index(filters["month"]) + 1
        conditions += f" AND MONTH(tss.posting_date) = {month}"
    
    if filters.get("year"):
        conditions += f" AND YEAR(tss.posting_date) = {filters.get('year')}"
    
    if filters.get("employee"):
        conditions += " AND tss.employee = '{}'".format(filters.get("employee"))
    
    return conditions


@frappe.whitelist()
def get_excel_data(filters):
	filters = json.loads(filters)
	# report_data = json.loads(data)
	file_name = "Salary Sheet"
	if filters.get("company"):
		file_name  += " - "+ filters.get("company") 
	data = ""

	return {
        "content": data,
        "filename": file_name,  
        "extension": 'xlsx'
    }


@frappe.whitelist()
def download_file():
	# workbook = openpyxl.Workbook(write_only=True)
	workbook = openpyxl.Workbook()
	current_sheet = workbook.active
	border_style_thin = Side(border_style='thin')
	border_style_thick = Side(border_style='thick')
	bold_font = Font(bold=True)

	# current_sheet.title = "Salary Sheet"

	# first row
	current_sheet.merge_cells("A1:D1")
	current_sheet['A1'].border = set_border(border_style_thick)
	current_sheet['D1'].border = set_border(border_style_thick)
	current_sheet.merge_cells("E1:N1")
    # 
	current_sheet.row_dimensions[1].height = 25
	current_sheet["E1"].fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
	update_border_font_align(current_sheet, 'E1', "Earning", border_style_thick); current_sheet['E1'].font = Font(size=14, bold=True)
	# current_sheet['E1'] = "Earning"
	# current_sheet['E1'].font = bold_font
	# current_sheet['E1'].alignment = Alignment(horizontal='center', vertical='center')
	# current_sheet['E1'].border = set_border(border_style_thick)
	current_sheet['N1'].border = set_border(border_style_thick)
	current_sheet.merge_cells("O1:P1")
	current_sheet['O1'].border = set_border(border_style_thick)
	current_sheet['P1'].border = set_border(border_style_thick)
	current_sheet.merge_cells("Q1:W1")
    #
	current_sheet["Q1"].fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
	update_border_font_align(current_sheet, 'Q1', "Deduction", border_style_thick) ; current_sheet['Q1'].font = Font(size=14, bold=True)
	# current_sheet['Q1'] = "Deduction"
	
	# current_sheet['Q1'].alignment = Alignment(horizontal='center', vertical='center')
	# current_sheet['Q1'].border = set_border(border_style_thick)
	current_sheet['W1'].border = set_border(border_style_thick)
      
	update_border_font_align(current_sheet, 'X1', "", border_style_thick)
	# current_sheet['Y1'] = "Total"
	# current_sheet['Y1'].font = bold_font
	# current_sheet['Y1'].border = set_border(border_style_thick)

	# second row
	# 
	current_sheet.row_dimensions[2].height = 50
	current_sheet.column_dimensions['B'].width = 18
	current_sheet.column_dimensions['D'].width = 15
	current_sheet.column_dimensions['F'].width = 15
	current_sheet.column_dimensions['G'].width = 15
	current_sheet.column_dimensions['H'].width = 15
	current_sheet.column_dimensions['I'].width = 15
	current_sheet.column_dimensions['K'].width = 15
	current_sheet.column_dimensions['L'].width = 15
	current_sheet.column_dimensions['M'].width = 15
	current_sheet.column_dimensions['Q'].width = 15
	current_sheet.column_dimensions['R'].width = 15
	current_sheet.column_dimensions['S'].width = 15
	current_sheet.column_dimensions['T'].width = 15
	current_sheet.column_dimensions['U'].width = 15
	update_border_font_align(current_sheet, 'A2', "S.No.", border_style_thin)
    # 
	update_border_font_align(current_sheet, 'B2', "Occupation", border_style_thin)
	update_border_font_align(current_sheet, 'C2', "Gross", border_style_thin)
	update_border_font_align(current_sheet, 'D2', "Employee", border_style_thin)
	update_border_font_align(current_sheet, 'E2', "Basic", border_style_thin)
	update_border_font_align(current_sheet, 'F2', "City Compensatory Allowance", border_style_thin)
	update_border_font_align(current_sheet, 'G2', "Washing Allowance", border_style_thin)
	update_border_font_align(current_sheet, 'H2', "House Rent Allowance	", border_style_thin)
	update_border_font_align(current_sheet, 'I2', "Overtime", border_style_thin)
	update_border_font_align(current_sheet, 'J2', "Arrear", border_style_thin)
	update_border_font_align(current_sheet, 'K2', "Incentive Pay", border_style_thin)
	update_border_font_align(current_sheet, 'L2', "Leave Encashment", border_style_thin)
	update_border_font_align(current_sheet, 'M2', "Travelling Allowance", border_style_thin)
	update_border_font_align(current_sheet, 'N2', "ABRY Scheme", border_style_thin)
	update_border_font_align(current_sheet, 'O2', "Earned Gross", border_style_thin)
	update_border_font_align(current_sheet, 'P2', "Net Gross", border_style_thin)
	update_border_font_align(current_sheet, 'Q2', "Employee Contribution PF", border_style_thin)
	update_border_font_align(current_sheet, 'R2', "Employee Contribution ESI", border_style_thin)
	update_border_font_align(current_sheet, 'S2', "Late Hours Deduction", border_style_thin)
	update_border_font_align(current_sheet, 'T2', "Tax Deducted at Source", border_style_thin)
	update_border_font_align(current_sheet, 'U2', "Professional Tax", border_style_thin)
	update_border_font_align(current_sheet, 'V2', "Income Tax", border_style_thin)
	update_border_font_align(current_sheet, 'W2', "GMI", border_style_thin)
	update_border_font_align(current_sheet, 'X2', "Total", border_style_thin)
	# current_sheet['A2'] = "S.No."
	# current_sheet['A2'].font = bold_font
	# current_sheet['A2'].border = set_border(border_style_thin)
	


	byte_file = io.BytesIO()
	workbook.save(byte_file)

	data = frappe._dict(frappe.local.form_dict)
	frappe.response["type"] = "binary"
	frappe.response["filecontent"] = byte_file.getvalue()
	frappe.response["filename"] = f"{data['filename']}.{data['extension']}"


def set_border(border_type):
	return Border(top=border_type, left=border_type, right=border_type, bottom=border_type)

def update_border_font_align(sheet, index, name, border_type) :
    sheet[index] = name
    sheet[index].font = Font(bold=True)
    sheet[index].border = set_border(border_type)
    # sheet[index].alignment = Alignment(horizontal='center', vertical='center', shrink_to_fit=True)
    sheet[index].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # sheet[index].width = 200
    return sheet

	# data = frappe._dict(frappe.local.form_dict)
	# frappe.response["filename"] = (
    #     frappe.scrub(f"{data['report_name']} {data['company']}") + ".txt"
    # )
	# frappe.response["filecontent"] = data["data"]
	# frappe.response["content_type"] = "application/txt"
	# frappe.response["type"] = "download"


      